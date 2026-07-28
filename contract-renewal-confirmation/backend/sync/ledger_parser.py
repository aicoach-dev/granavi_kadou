"""
Excel台帳パーサー — Lambda用

scripts/analyze_ledger.py の parse_ledger() を Lambda 向けに移植したもの。
引数が Path ではなく bytes になっている点が主な違い。
DynamoDB等のAWSリソースには接続しない。
"""
import io
import json
from datetime import date, datetime, timedelta


def load_config(config_path: str) -> dict:
    with open(config_path, encoding="utf-8") as f:
        return json.load(f)


def excel_serial_to_date(serial) -> str | None:
    """Excel の日付シリアル値（または datetime オブジェクト）を ISO 日付文字列に変換する。"""
    if serial is None:
        return None
    if isinstance(serial, datetime):
        return serial.date().isoformat()
    if isinstance(serial, date):
        return serial.isoformat()
    if isinstance(serial, (int, float)):
        return (date(1899, 12, 30) + timedelta(days=int(serial))).isoformat()
    return None


def normalize_company_name(raw_title: str, cfg: dict) -> str:
    norm = cfg.get("company_name_normalization", {})
    sep = norm.get("strip_suffix_after", "：")
    name = raw_title.split(sep)[0].strip()
    return norm.get("overrides", {}).get(name, name)


def normalize_consent(raw_value, mapping: dict) -> str | None:
    if raw_value is None:
        return None
    v = str(raw_value).strip()
    for status, values in mapping.items():
        if v in values:
            return status
    return f"unknown:{v}"


def parse_ledger(content: bytes, cfg: dict) -> list[dict]:
    """
    Excel台帳をバイト列から解析して本人データのリストを返す。
    scripts/analyze_ledger.py の parse_ledger() に対応する Lambda 版。
    """
    import openpyxl  # Lambda パッケージ内でのみ利用

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    sheet_name = cfg.get("sheet_name")
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    title_col = cfg["block_detection"]["title_col_index"]
    header_marker_col = cfg["header_row_detection"]["col_index"]
    header_marker = cfg["header_row_detection"]["marker"]
    name_col = cfg["data_row_detection"]["name_col_index"]
    col_cfg = cfg["columns"]
    consent_map = cfg.get("consent_value_mapping", {})
    skip_companies = set(cfg.get("skip_blocks", {}).get("companies", []))
    block_overrides = cfg.get("block_overrides", {})

    rows = list(ws.iter_rows(values_only=True))
    current_block_company = None
    in_data_zone = False
    results = []

    for row in rows:
        title_cell = row[title_col]
        if title_cell is not None and isinstance(title_cell, str):
            current_block_company = normalize_company_name(title_cell.strip(), cfg)
            in_data_zone = False
            continue

        if current_block_company is None or current_block_company in skip_companies:
            continue

        header_cell = row[header_marker_col]
        if (
            header_cell is not None
            and isinstance(header_cell, str)
            and header_marker in header_cell
        ):
            in_data_zone = True
            continue

        if not in_data_zone:
            continue

        name_cell = row[name_col]
        if name_cell is None or not isinstance(name_cell, str):
            continue

        override = block_overrides.get(current_block_company, {})
        company_raw = row[col_cfg["company"]["col_index"]]
        company = str(company_raw).strip() if company_raw else current_block_company

        p_start_col = override.get("period_start_col_index", col_cfg["period_start"]["col_index"])
        p_end_col = override.get("period_end_col_index", col_cfg["period_end"]["col_index"])

        period_start_raw = row[p_start_col] if p_start_col is not None else None
        period_end_raw = row[p_end_col]
        consent_raw = row[col_cfg["consent"]["col_index"]]

        results.append(
            {
                "block_company": current_block_company,
                "company": company,
                "name": str(name_cell).strip(),
                "contract_type": row[col_cfg["contract_type"]["col_index"]],
                "period_start": excel_serial_to_date(period_start_raw),
                "period_end": excel_serial_to_date(period_end_raw),
                "consent_raw": consent_raw,
                "consent": normalize_consent(consent_raw, consent_map),
            }
        )

    return results

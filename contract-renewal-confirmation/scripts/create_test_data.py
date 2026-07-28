"""
テストデータ生成スクリプト（ローカル実行用）

生成するファイル:
  1. synthetic.xlsx  — 6つの完了条件を各1件で網羅した合成テストファイル（7行）
  2. sanitized.xlsx  — 実データ（scripts/sample-data/）の氏名を仮名化したファイル
                      実データが存在しない場合はスキップ

生成後、S3バケット SYNC_DATA_BUCKET に以下のキーでアップロードする:
  test-data/synthetic.xlsx
  test-data/sanitized.xlsx （実データがある場合のみ）

Usage:
  python scripts/create_test_data.py

環境変数:
  SYNC_DATA_BUCKET  アップロード先バケット名（省略時は下記デフォルト）

注: Lambda の EXCEL_KEY 環境変数を変更することで、
  synthetic と sanitized を切り替えてテストできる。
"""
import io
import json
import logging
import os
from datetime import date
from pathlib import Path

import boto3
import openpyxl
from openpyxl import Workbook

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

SCRIPT_DIR = Path(__file__).parent
CONFIG_PATH = SCRIPT_DIR / "config" / "ledger_mapping.json"
SAMPLE_DATA_DIR = SCRIPT_DIR / "sample-data"
SYNC_DATA_BUCKET = os.environ.get(
    "SYNC_DATA_BUCKET",
    "contract-renewal-sync-data-698212246219-ap-northeast-1",
)


def date_to_serial(d: date) -> int:
    """Python date を Excel シリアル値に変換する。"""
    return (d - date(1899, 12, 30)).days


def load_config() -> dict:
    with open(CONFIG_PATH, encoding="utf-8") as f:
        return json.load(f)


# ============================================================
# 合成テストファイルの生成
# ============================================================

# 各タプル: (block_company, name, contract_type, period_start, period_end, consent_value)
# 2026-07-28 実行時の期待結果:
#   ケース1: 候補作成 + opsConsentResult='consent' プリセット（本人了承済み）
#   ケース2: 候補作成 + opsConsentResult='consent' プリセット（了承みなし）
#   ケース3: 候補作成（consent=unknown:謎の値2026 として記録、通常候補）
#   ケース4: 候補作成（consent=null、通常候補）
#   ケース5: 候補作成なし（基準日前: 2026-08-17 が基準日 < 2026-07-28）
#   ケース6: 候補作成なし（委託・許可リスト外）
#   ケース7: 候補作成（委託・許可リスト内 = イーソル株式会社）
SYNTHETIC_CASES = [
    (
        "テスト派遣株式会社",
        "テスト氏名1",
        "派遣",
        date(2026, 7, 1),
        date(2026, 9, 30),
        "本人了承済み",
    ),
    (
        "テスト派遣株式会社",
        "テスト氏名2",
        "派遣",
        date(2026, 7, 1),
        date(2026, 9, 30),
        "了承みなし",
    ),
    (
        "テスト派遣株式会社",
        "テスト氏名3",
        "派遣",
        date(2026, 7, 1),
        date(2026, 9, 30),
        "謎の値2026",
    ),
    (
        "テスト派遣株式会社",
        "テスト氏名4",
        "派遣",
        date(2026, 7, 1),
        date(2026, 9, 30),
        None,
    ),
    (
        "テスト派遣株式会社",
        "テスト氏名5",
        "派遣",
        date(2026, 10, 1),
        date(2026, 12, 31),
        None,
    ),
    (
        "テスト委託除外株式会社",
        "テスト氏名6",
        "委託",
        date(2026, 7, 1),
        date(2026, 9, 30),
        None,
    ),
    (
        "イーソル株式会社",
        "テスト氏名7",
        "委託",
        date(2026, 7, 1),
        date(2026, 9, 30),
        None,
    ),
]


def create_synthetic_workbook() -> Workbook:
    """6つの完了条件を各1件ずつ網羅した合成テスト台帳を生成する。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "台帳一覧2026年度以降（進行中）"

    # ブロックごとにグループ化（順序維持）
    blocks: dict[str, list] = {}
    for case in SYNTHETIC_CASES:
        block = case[0]
        if block not in blocks:
            blocks[block] = []
        blocks[block].append(case)

    row_idx = 1
    for block_company, cases in blocks.items():
        # ブロックタイトル行（A列のみ）
        ws.cell(row=row_idx, column=1, value=block_company)
        row_idx += 1

        # ヘッダー行
        headers = [
            "",
            "手続き",
            "客先名",
            "氏名",
            "形態",
            "期間（年度）",
            "開始日",
            "終了日",
            "契約延長連絡",
            "本人確認",
        ]
        for col, h in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col, value=h)
        row_idx += 1

        # データ行
        for _, name, ctype, pstart, pend, consent_val in cases:
            ws.cell(row=row_idx, column=2, value="進行中")
            ws.cell(row=row_idx, column=3, value=block_company)
            ws.cell(row=row_idx, column=4, value=name)
            ws.cell(row=row_idx, column=5, value=ctype)
            ws.cell(row=row_idx, column=6, value=f"{pstart.year}年度")
            ws.cell(row=row_idx, column=7, value=date_to_serial(pstart))
            ws.cell(row=row_idx, column=8, value=date_to_serial(pend))
            if consent_val is not None:
                ws.cell(row=row_idx, column=10, value=consent_val)
            row_idx += 1

    return wb


# ============================================================
# 実データの氏名仮名化
# ============================================================


def anonymize_workbook(source_path: Path, cfg: dict) -> Workbook | None:
    """
    実データExcelの氏名列（D列 = name_col_index=3）を仮名に置換して返す。
    客先名・期間・種別・本人確認値はそのまま維持する。
    """
    if not source_path.exists():
        logger.info(f"実データが見つかりません（スキップ）: {source_path}")
        return None

    logger.info(f"実データを読み込み中: {source_path}")
    wb = openpyxl.load_workbook(source_path)

    sheet_name = cfg.get("sheet_name")
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    title_col_idx = cfg["block_detection"]["title_col_index"]        # 0-based
    header_marker_col_idx = cfg["header_row_detection"]["col_index"]  # 0-based
    header_marker = cfg["header_row_detection"]["marker"]
    name_col_idx = cfg["data_row_detection"]["name_col_index"]        # 0-based

    in_data_zone = False
    anon_counter = 1

    for row in ws.iter_rows():
        title_cell = row[title_col_idx]
        if title_cell.value is not None and isinstance(title_cell.value, str):
            in_data_zone = False
            continue

        header_cell = row[header_marker_col_idx]
        if (
            header_cell.value is not None
            and isinstance(header_cell.value, str)
            and header_marker in header_cell.value
        ):
            in_data_zone = True
            continue

        if not in_data_zone:
            continue

        name_cell = row[name_col_idx]
        if name_cell.value is not None and isinstance(name_cell.value, str):
            name_cell.value = f"仮名テスト氏名{anon_counter:03d}"
            anon_counter += 1

    logger.info(f"仮名化完了: {anon_counter - 1} 件の氏名を置換しました")
    return wb


# ============================================================
# S3アップロード
# ============================================================


def upload_workbook(wb: Workbook, bucket: str, key: str) -> None:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    s3 = boto3.client("s3")
    s3.put_object(Bucket=bucket, Key=key, Body=buf.read())
    logger.info(f"S3アップロード完了: s3://{bucket}/{key}")


# ============================================================
# メイン
# ============================================================


def main() -> None:
    cfg = load_config()

    logger.info("=== 合成テストファイルの生成 ===")
    synthetic_wb = create_synthetic_workbook()
    upload_workbook(synthetic_wb, SYNC_DATA_BUCKET, "test-data/synthetic.xlsx")

    logger.info("合成テスト ケース一覧:")
    labels = [
        "候補+opsConsentPreset（本人了承済み）",
        "候補+opsConsentPreset（了承みなし）",
        "候補+unknown値（通常フロー）",
        "候補+consent=null（通常フロー）",
        "候補作成なし（基準日前: 10/1開始）",
        "候補作成なし（委託・許可リスト外）",
        "候補（委託・許可リスト内: イーソル）",
    ]
    for i, (case, label) in enumerate(zip(SYNTHETIC_CASES, labels), 1):
        _, name, ctype, pstart, _, cv = case
        logger.info(f"  ケース{i}: {name} / {ctype} / start={pstart} / consent={cv!r} → {label}")

    logger.info("=== 実データの氏名仮名化 ===")
    sample_files = sorted(SAMPLE_DATA_DIR.glob("*.xlsx")) if SAMPLE_DATA_DIR.exists() else []
    if sample_files:
        sample_path = sample_files[0]
        logger.info(f"対象ファイル: {sample_path.name}")
        sanitized_wb = anonymize_workbook(sample_path, cfg)
        if sanitized_wb:
            upload_workbook(sanitized_wb, SYNC_DATA_BUCKET, "test-data/sanitized.xlsx")
    else:
        logger.info(f"実データが {SAMPLE_DATA_DIR} に見つかりません（sanitized.xlsx はスキップ）")

    logger.info("=== テストデータ生成・アップロード完了 ===")


if __name__ == "__main__":
    main()

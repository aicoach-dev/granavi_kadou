"""
Excel台帳 技術検証スクリプト（Round 2 スパイク用）
・本番コードではない。週次同期Lambdaの実装方針確認が目的。
・scripts/config/ledger_mapping.json を外部設定として使用する。
・DynamoDB 等の AWS リソースには一切接続しない。
"""

import openpyxl
import json
import sys
from datetime import date, timedelta
from pathlib import Path

SCRIPT_DIR = Path(__file__).parent
CONFIG_PATH = SCRIPT_DIR / "config" / "ledger_mapping.json"
SAMPLE_PATH = SCRIPT_DIR / "sample-data" / "グラビティ契約更新台帳サンプル.xlsx"


def load_config(path: Path) -> dict:
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def excel_serial_to_date(serial) -> str | None:
    """Excel の日付シリアル値を ISO 日付文字列に変換する。"""
    if serial is None or not isinstance(serial, (int, float)):
        return None
    return (date(1899, 12, 30) + timedelta(days=int(serial))).isoformat()


def normalize_company_name(raw_title: str, cfg: dict) -> str:
    """ブロックタイトル（A列）の社名を正規化する。"""
    norm = cfg.get("company_name_normalization", {})
    # 「：」以降を除去（注釈付きタイトルへの対応）
    sep = norm.get("strip_suffix_after", "：")
    name = raw_title.split(sep)[0].strip()
    # overridesで上書き
    return norm.get("overrides", {}).get(name, name)


def normalize_consent(raw_value, mapping: dict) -> str | None:
    """本人確認列の値を confirmed / pending / declined / null に正規化する。"""
    if raw_value is None:
        return None
    v = str(raw_value).strip()
    for status, values in mapping.items():
        if v in values:
            return status
    return f"unknown:{v}"


def parse_ledger(wb_path: Path, cfg: dict) -> list[dict]:
    """
    Excel台帳を解析して本人データのリストを返す。
    各要素は必須カラムのみを持つ辞書。
    """
    wb = openpyxl.load_workbook(wb_path, data_only=True)
    sheet_name = cfg.get("sheet_name")
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    title_col = cfg["block_detection"]["title_col_index"]
    header_marker_col = cfg["header_row_detection"]["col_index"]
    header_marker = cfg["header_row_detection"]["marker"]
    name_col = cfg["data_row_detection"]["name_col_index"]
    col_cfg = cfg["columns"]
    consent_map = cfg.get("consent_value_mapping", {})
    skip_companies = set(cfg.get("skip_blocks", {}).get("companies", []))
    block_overrides = cfg.get("block_overrides", {})

    rows = list(ws.iter_rows(values_only=True))

    current_block_company = None
    in_data_zone = False  # ヘッダー行通過後フラグ
    results = []

    for i, row in enumerate(rows):
        # ---- ブロック開始判定 ----
        title_cell = row[title_col]
        if title_cell is not None and isinstance(title_cell, str):
            raw_title = title_cell.strip()
            current_block_company = normalize_company_name(raw_title, cfg)
            in_data_zone = False
            continue

        if current_block_company is None:
            continue
        if current_block_company in skip_companies:
            continue

        # ---- ヘッダー行判定 ----
        header_cell = row[header_marker_col]
        if header_cell is not None and isinstance(header_cell, str) and header_marker in header_cell:
            in_data_zone = True
            continue

        if not in_data_zone:
            continue

        # ---- データ行判定 ----
        name_cell = row[name_col]
        if name_cell is None or not isinstance(name_cell, str):
            continue

        # ブロックオーバーライドを適用
        override = block_overrides.get(current_block_company, {})

        # 各フィールド抽出
        company_raw = row[col_cfg["company"]["col_index"]]
        # データ行のC列が省略名の場合、ブロック正規化名をフォールバックとして使う
        company = str(company_raw).strip() if company_raw else current_block_company

        p_start_col = override.get("period_start_col_index", col_cfg["period_start"]["col_index"])
        p_end_col = override.get("period_end_col_index", col_cfg["period_end"]["col_index"])

        period_start_raw = row[p_start_col] if p_start_col is not None else None
        period_end_raw = row[p_end_col]

        record = {
            "block_company": current_block_company,
            "company": company,
            "name": str(name_cell).strip(),
            "contract_type": row[col_cfg["contract_type"]["col_index"]],
            "period_start": excel_serial_to_date(period_start_raw),
            "period_end": excel_serial_to_date(period_end_raw),
            "consent_raw": row[col_cfg["consent"]["col_index"]],
            "consent": normalize_consent(row[col_cfg["consent"]["col_index"]], consent_map),
        }
        results.append(record)

    return results


def print_report(records: list[dict]) -> None:
    print(f"=== 抽出結果: {len(records)} 件 ===\n")

    # ブロック別集計
    block_counts: dict[str, int] = {}
    for r in records:
        block_counts[r["block_company"]] = block_counts.get(r["block_company"], 0) + 1

    print("--- ブロック別件数 ---")
    for company, cnt in block_counts.items():
        print(f"  {company}: {cnt}件")

    # 本人確認値の分布
    print("\n--- 本人確認 正規化後の分布 ---")
    consent_counts: dict[str, int] = {}
    for r in records:
        v = r["consent"] or "(null)"
        consent_counts[v] = consent_counts.get(v, 0) + 1
    for v, cnt in sorted(consent_counts.items(), key=lambda x: -x[1]):
        print(f"  {v}: {cnt}件")

    # 契約形態の分布
    print("\n--- 契約形態の分布 ---")
    type_counts: dict[str, int] = {}
    for r in records:
        t = str(r["contract_type"]) if r["contract_type"] else "(null)"
        type_counts[t] = type_counts.get(t, 0) + 1
    for t, cnt in sorted(type_counts.items(), key=lambda x: -x[1]):
        print(f"  {t}: {cnt}件")

    # 日付変換の確認
    date_parse_failures = [r for r in records if r["period_start"] is None and r.get("period_end") is None]
    print(f"\n--- 期間日付が両方nullの行: {len(date_parse_failures)}件 ---")
    for r in date_parse_failures:
        print(f"  {r['name']} ({r['block_company']}): start=None, end=None")

    # "unknown:" が含まれる consent 値（表記ゆれ候補）
    unknowns = [r for r in records if r["consent"] and r["consent"].startswith("unknown:")]
    print(f"\n--- 未知の本人確認値（要設定追加）: {len(unknowns)}件 ---")
    for r in unknowns:
        print(f"  {r['name']} ({r['block_company']}): {r['consent']}")

    # サンプル出力（先頭5件）
    print("\n--- サンプル（先頭5件） ---")
    for r in records[:5]:
        print(f"  {r['name']} / {r['company']} / {r['contract_type']} / "
              f"{r['period_start']}〜{r['period_end']} / consent={r['consent']}")


if __name__ == "__main__":
    wb_path = Path(sys.argv[1]) if len(sys.argv) > 1 else SAMPLE_PATH
    cfg = load_config(CONFIG_PATH)
    records = parse_ledger(wb_path, cfg)
    print_report(records)

    # JSON として出力（オプション）
    out_path = wb_path.parent / "extracted_records.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)
    print(f"\n→ 抽出結果を {out_path} に保存しました。")
